import openpyxl

path = './プラモデル・鉄道模型在庫リスト260416 (2).xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print('Sheets:', wb.sheetnames)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print('=== ' + sheet_name + ' ===')
    print('Max row:', ws.max_row, 'Max col:', ws.max_column)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 3: break
        print('Row', i+1, ':', list(row)[:12])