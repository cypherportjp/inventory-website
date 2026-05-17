import openpyxl, os
path = os.path.join(os.path.dirname(__file__), 'プラモデル・鉄道模型在庫リスト260416 (2).xlsx')
wb = openpyxl.load_workbook(path, data_only=True)
print('Sheets:', wb.sheetnames)
for s in wb.sheetnames:
    ws = wb[s]
    print('=== ' + s + ' === Rows:', ws.max_row, 'Cols:', ws.max_column)
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 2: break
        print(i+1, list(row)[:12])